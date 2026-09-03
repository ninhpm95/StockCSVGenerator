"""
Reads and overwrites files in INPUT_FOLDER in place. No xlsx->csv
conversion; each file keeps its original format. Only files that
actually get trimmed are logged in detail; everything else is folded
into a single "kept as is" count.

For .csv files:
    - Scan column A for any string in SEARCH_STRS. On the row where a
      string hits its configured occurrence count, delete that row and
      everything after it.

For .xlsx files:
    - Find the first sheet name in SHEET_NAMES that exists in the
      workbook (other sheets are left alone). Scan that sheet's column A
      the same way and delete rows from the match onward.
    - If no matching sheet name exists, the file is left untouched.

Requires: pip install openpyxl
"""

import csv
import io
from pathlib import Path
import openpyxl
from .constants import INPUT_FOLDER, SHEET_NAMES, SEARCH_STRS


def _normalize(s):
    """Strip leading/trailing whitespace, including non-breaking (\\xa0)
    and full-width (\\u3000) spaces that Excel/CSV exports sometimes embed
    but that look identical to a normal space or nothing at all."""
    return s.replace("\xa0", " ").replace("\u3000", " ").strip()


def find_cutoff(rows):
    """Scan rows top-down, tracking how many times each search string has
    matched column A (after whitespace normalization). Return
    (row_index, search_str, target_count) for the first search string to
    reach its configured occurrence count, or None if none ever does."""
    counts = {search_str: 0 for search_str, _ in SEARCH_STRS}
    normalized_targets = {search_str: _normalize(search_str) for search_str, _ in SEARCH_STRS}
    for i, row in enumerate(rows):
        first_col = _normalize(str(row[0])) if row and row[0] is not None else ""
        for search_str, target_count in SEARCH_STRS:
            if first_col == normalized_targets[search_str]:
                counts[search_str] += 1
                if counts[search_str] == target_count:
                    return i, search_str, target_count
    return None


def _near_misses(rows, search_strs):
    """Diagnostic: cells that loosely resemble a search string
    (case-insensitive substring) but didn't exact-match after
    normalization. Surfaces hidden-character/casing mismatches instead
    of failing silently."""
    normalized_targets = {_normalize(s) for s in search_strs}
    found = []
    for i, row in enumerate(rows):
        raw = str(row[0]) if row and row[0] is not None else ""
        normalized = _normalize(raw)
        if normalized in normalized_targets or not normalized:
            continue
        for s in search_strs:
            if s.lower() in normalized.lower():
                found.append((i, repr(raw)))
                break
    return found


def _read_csv_rows(csv_path: Path):
    """Try utf-8 (with or without BOM) first; fall back to cp932
    (Shift-JIS) for Japanese exports that aren't UTF-8. Returns
    (rows, encoding_used), where encoding_used is 'utf-8-sig' only if
    the file actually had a BOM, so writes can preserve that instead of
    always adding one."""
    raw = csv_path.read_bytes()
    has_bom = raw.startswith(b"\xef\xbb\xbf")
    try:
        text = raw.decode("utf-8-sig" if has_bom else "utf-8")
        return list(csv.reader(io.StringIO(text, newline=""))), ("utf-8-sig" if has_bom else "utf-8")
    except UnicodeDecodeError:
        pass
    try:
        text = raw.decode("cp932")
        return list(csv.reader(io.StringIO(text, newline=""))), "cp932"
    except UnicodeDecodeError:
        pass
    raise UnicodeDecodeError(
        "unknown", b"", 0, 1, f"Could not decode {csv_path.name} as utf-8(-sig) or cp932"
    )


def process_csv(csv_path: Path):
    """Returns (change_detail_or_None, near_misses)."""
    rows, encoding = _read_csv_rows(csv_path)

    match = find_cutoff(rows)
    if match is None:
        return None, _near_misses(rows, [s for s, _ in SEARCH_STRS])
    cutoff, search_str, target_count = match

    rows = rows[:cutoff]
    with open(csv_path, "w", newline="", encoding=encoding) as f:
        csv.writer(f).writerows(rows)

    return {
        "name": csv_path.name,
        "search_str": search_str,
        "target_count": target_count,
        "row": cutoff + 1,
    }, []


def _unmerge_from_row(ws, start_row):
    """Unmerge any merged cell range that overlaps or lies entirely
    within [start_row, end of sheet]. These ranges are either about to
    be deleted wholesale or straddle the cutoff; either way, leaving
    them merged across a delete_rows() call can corrupt the merge
    metadata (openpyxl doesn't rewrite merged ranges when rows are
    deleted). Ranges entirely above start_row are left untouched."""
    to_unmerge = [
        str(merged_range)
        for merged_range in list(ws.merged_cells.ranges)
        if merged_range.max_row >= start_row
    ]
    for merged_range in to_unmerge:
        ws.unmerge_cells(merged_range)


def process_xlsx(xlsx_path: Path):
    """Returns (change_detail_or_None, near_misses)."""
    wb = openpyxl.load_workbook(xlsx_path)

    target_sheet = None
    for name in SHEET_NAMES:
        if name in wb.sheetnames:
            target_sheet = name
            break

    if target_sheet is None:
        return None, []

    ws = wb[target_sheet]
    rows = list(ws.iter_rows(values_only=True))
    match = find_cutoff(rows)

    if match is None:
        return None, _near_misses(rows, [s for s, _ in SEARCH_STRS])
    cutoff, search_str, target_count = match

    # Unmerge any merged ranges touching the rows we're about to remove,
    # then delete bottom-up so row numbers don't shift mid-delete.
    _unmerge_from_row(ws, cutoff + 1)
    ws.delete_rows(cutoff + 1, ws.max_row - cutoff)
    wb.save(xlsx_path)

    return {
        "name": xlsx_path.name,
        "search_str": search_str,
        "target_count": target_count,
        "row": cutoff + 1,
        "sheet": target_sheet,
    }, []


def run():
    paths = [p for p in sorted(INPUT_FOLDER.iterdir()) if p.suffix.lower() in (".csv", ".xlsx")]
    print(f"Scanning {len(paths)} file(s) in {INPUT_FOLDER}")

    changed = []
    failures = []
    near_miss_report = []  # (filename, [(row_index, repr)])

    for path in paths:
        try:
            if path.suffix.lower() == ".csv":
                result, misses = process_csv(path)
            else:
                result, misses = process_xlsx(path)
            if result is not None:
                changed.append(result)
            elif misses:
                near_miss_report.append((path.name, misses))
        except Exception as e:
            failures.append((path.name, str(e)))

    kept_as_is = len(paths) - len(changed) - len(failures)
    print(f"Num of files kept as is: {kept_as_is}")

    if changed:
        print("Changed files:")
        for c in changed:
            sheet_note = f" (sheet '{c['sheet']}')" if "sheet" in c else ""
            print(f"  {c['name']}: \"{c['search_str']}\" ({c['target_count']}) found on row {c['row']}{sheet_note}")

    if failures:
        print("Errored files:")
        for name, err in failures:
            print(f"  {name}: {err}")

    if near_miss_report:
        print("Near-misses (cell resembles a search string but didn't exact-match):")
        for name, misses in near_miss_report:
            for row_index, raw in misses:
                print(f"  {name} row {row_index + 1}: {raw}")
